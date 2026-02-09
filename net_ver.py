# quantum_server_gui_net.py
# сервер: читает Quantum (COM/файл), отображает, раздаёт state клиентам по TCP (JSON lines)
# + загрузка Excel (.xlsx): A=bib, B=name
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import argparse
import csv
import json
import queue
import re
import os
from urllib.parse import unquote
import socket
import threading
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from typing import Optional, List, Dict, Any

try:
    import serial
    from serial.tools import list_ports
except Exception:
    serial = None
    list_ports = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


FREEZE_SEC = 2.0
FLAG_NAME_RE = re.compile(r"^[A-Za-z0-9_\-\.]{1,64}\.(?:png|PNG)$")

MSG_RE = re.compile(r"(?:[A-Z]{0,64})?(DN|DA|DS|DI|DF)\|")
TIME_HMS_RE = re.compile(r"(\d{2}):(\d{2}):(\d{2})\.(\d{3})")
TIME_MS_RE = re.compile(r"(\d+):(\d{2})\.(\d{3})")

FLAG_MAP = {
    "мск": "moscow.png",
    "москва": "moscow.png",
    "спб": "spb.png",
    "санкт-петербург": "spb.png",
    "омск": "omsk.png",
    "хабаровск": "khabarovsk.png",
    "беларусь": "belarus.png",
    "иркутск": "irkutsk.png",
    "тула": "tula.png",
}

def _norm_flag_key(x) -> str:
    if x is None:
        return ""
    s = str(x).strip().lower().replace("ё", "е")
    s = re.sub(r"\s+", "", s)
    return s

def flag_for_excel_value(x) -> str:
    k = _norm_flag_key(x)
    return FLAG_MAP.get(k, "")


import os, tempfile, sys

def resource_path(rel: str) -> str:
    if getattr(sys, "frozen", False):
        return os.path.join(sys._MEIPASS, rel)   # из exe (временная распаковка)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), rel)

class OverlayHttp:
    def __init__(self, app, host="0.0.0.0", port=8099, overlay_html_path=None, flags_dir=None):
        self.app = app
        self.host = host
        self.port = port
        self.overlay_html_path = overlay_html_path
        self.flags_dir = flags_dir
        self.httpd = None
        self.thr = None


    def start(self):
        app = self.app
        overlay_html_path = self.overlay_html_path
        flags_dir = self.flags_dir  # <<<

        class H(BaseHTTPRequestHandler):
            def _send(self, code, ctype, data: bytes):
                self.send_response(code)
                self.send_header("Content-Type", ctype)
                self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
                self.end_headers()
                self.wfile.write(data)

            def do_GET(self):
                path = self.path.split("?", 1)[0]

                if path in ("/", "/overlay_test.html"):
                    if not overlay_html_path or not os.path.isfile(overlay_html_path):
                        return self._send(404, "text/plain; charset=utf-8", b"overlay.html not found")
                    try:
                        with open(overlay_html_path, "rb") as f:
                            data = f.read()
                        return self._send(200, "text/html; charset=utf-8", data)
                    except Exception:
                        return self._send(500, "text/plain; charset=utf-8", b"cannot read overlay.html")
                if path.startswith("/flags/"):
                    if not flags_dir:
                        return self._send(404, "text/plain; charset=utf-8", b"flags dir not set")

                    fn = os.path.basename(unquote(path[len("/flags/"):]))
                    if not FLAG_NAME_RE.match(fn):
                        return self._send(404, "text/plain; charset=utf-8", b"bad flag name")

                    fp = os.path.join(flags_dir, fn)
                    if not os.path.isfile(fp):
                        return self._send(404, "text/plain; charset=utf-8", b"flag not found")

                    try:
                        with open(fp, "rb") as f:
                            data = f.read()
                        return self._send(200, "image/png", data)
                    except Exception:
                        return self._send(500, "text/plain; charset=utf-8", b"cannot read flag")

                if self.path.startswith("/state.json"):
                    try:
                        with app._obs_lock:
                            payload = dict(app._obs_payload)
                        data = (json.dumps(payload, ensure_ascii=False) + "\n").encode("utf-8")
                        return self._send(200, "application/json; charset=utf-8", data)
                    except Exception:
                        return self._send(500, "application/json; charset=utf-8", b"{}")

                return self._send(404, "text/plain; charset=utf-8", b"not found")

            def log_message(self, format, *args):
                return

        try:
            self.httpd = ThreadingHTTPServer((self.host, self.port), H)
        except Exception as e:
            try:
                app.q.put({"kind": "err", "data": f"OVERLAY HTTP: не стартанул на {self.host}:{self.port} ({e})"})
            except Exception:
                pass
            self.httpd = None
            return

        self.thr = threading.Thread(target=self.httpd.serve_forever, kwargs={"poll_interval": 0.2}, daemon=True)
        self.thr.start()

        try:
            app.q.put({"kind": "evt", "data": {"type": "other",
                                               "raw": f"OVERLAY HTTP: запущен http://{self.host}:{self.port}/overlay_test.html"}})
        except Exception:
            pass

    def stop(self):
        try:
            if self.httpd:
                self.httpd.shutdown()
                self.httpd.server_close()
        except Exception:
            pass


def _atomic_write_text(path: str, text: str):
    d = os.path.dirname(path) or "."
    os.makedirs(d, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=".tmp_", dir=d, text=True)
    try:
        with os.fdopen(fd, "w", encoding="utf-8", newline="\n") as f:
            f.write(text)
        os.replace(tmp, path)
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass
def _mono_to_epoch(mono_point: float) -> float:
    return time.time() + (mono_point - time.monotonic())


def _fmt_obs_two(b1: str, n1: str, t1: str, b2: str, n2: str, t2: str) -> str:
    left = (f"{b1} {n1}".strip() + "\n" + (t1 or "")).strip()
    right = (f"{b2} {n2}".strip() + "\n" + (t2 or "")).strip()

    # две “колонки” текстом; OBS сам рисует моношириной/выравниванием как настроишь
    # если второго нет — просто пустая правая часть
    return left + "\n\n" + right + "\n"


def strip_ctrl(s: str) -> str:
    if not s:
        return ""
    return "".join(ch for ch in s if ch.isprintable() or ch in "\t \n\r")


def clean_token(s: str) -> str:
    return strip_ctrl(s).replace("\r", "").replace("\n", "").strip()


def extract_first_int(token: Optional[str]) -> Optional[str]:
    if token is None:
        return None
    t = clean_token(token)
    m = re.search(r"(\d+)", t)
    if not m:
        return None
    try:
        n = int(m.group(1))
    except Exception:
        return None
    if n == 0:
        return None
    return str(n)


def parse_time_any(token: Optional[str]) -> Optional[float]:
    if token is None:
        return None
    s = clean_token(token)
    if not s:
        return None

    m = TIME_HMS_RE.search(s)
    if m:
        hh, mm, ss, ms = map(int, m.groups())
        return hh * 3600 + mm * 60 + ss + ms / 1000.0

    m = TIME_MS_RE.search(s)
    if m:
        mm, ss, ms = map(int, m.groups())
        return mm * 60 + ss + ms / 1000.0

    m = re.search(r"[+-]?\d+\.\d+", s)
    if m:
        try:
            return float(m.group(0))
        except Exception:
            return None

    m = re.search(r"[+-]?\d+", s)
    if m:
        try:
            return float(m.group(0))
        except Exception:
            return None

    return None


def find_primary_time(tokens: List[str]) -> Optional[float]:
    for t in tokens:
        v = parse_time_any(t)
        if v is not None:
            return v
    return None


def fmt_time(sec: Optional[float]) -> str:
    if sec is None:
        return ""
    sign = "-" if sec < 0 else ""
    sec = abs(sec)
    total_ms = int(round(sec * 1000))
    s = (total_ms // 1000) % 60
    m = (total_ms // 60000) % 60
    h = total_ms // 3600000
    ms = total_ms % 1000
    if h > 0:
        return f"{sign}{h:d}:{m:02d}:{s:02d}.{ms:03d}"
    if total_ms >= 60000:
        return f"{sign}{m:d}:{s:02d}.{ms:03d}"
    return f"{sign}{total_ms/1000:.3f}"


def fmt_live(sec: Optional[float]) -> str:
    if sec is None:
        return "—"
    total_ms = int(max(0, round(sec * 1000)))
    s = (total_ms // 1000) % 60
    m = (total_ms // 60000) % 60
    h = total_ms // 3600000
    ms = total_ms % 1000
    if h > 0:
        return f"{h:d}:{m:02d}:{s:02d}.{ms:03d}"
    if total_ms >= 60000:
        return f"{m:d}:{s:02d}.{ms:03d}"
    return f"{total_ms/1000:.3f}"


def split_sort_key(x: str):
    x = str(x).strip()
    if x.isdigit():
        return (0, int(x))
    return (1, x)


def safe_col_id(x: str) -> str:
    x = str(x)
    out = []
    for ch in x:
        out.append(ch if ch.isalnum() else "_")
    return "s_" + "".join(out)


def split_stream(buffer: str):
    if not buffer:
        return [], ""

    msgs = []
    matches = list(MSG_RE.finditer(buffer))

    if matches:
        first = matches[0].start()
        if first > 0:
            buffer = buffer[first:]
            matches = list(MSG_RE.finditer(buffer))

        if len(matches) >= 2:
            for i in range(len(matches) - 1):
                a = matches[i].start()
                b = matches[i + 1].start()
                msgs.append(buffer[a:b])
            rest = buffer[matches[-1].start():]
            return msgs, rest

        start = matches[0].start()
        tail = buffer[start:]

        m_end = re.search(r"S\d{1,2}", tail)
        if m_end:
            endpos = start + m_end.end()
            msgs.append(buffer[start:endpos])
            rest = buffer[endpos:]
            return msgs, rest

        nl = buffer.find("\n", start)
        if nl != -1:
            msgs.append(buffer[start:nl + 1])
            return msgs, buffer[nl + 1:]

        return [], buffer

    if "\n" in buffer:
        parts = buffer.splitlines(True)
        for p in parts[:-1]:
            msgs.append(p)
        return msgs, parts[-1]

    return [], buffer


def parse_message(raw: str) -> Optional[Dict[str, Any]]:
    if not raw:
        return None

    raw_clean = strip_ctrl(raw).strip()
    if not raw_clean:
        return None

    m = MSG_RE.search(raw_clean)
    if not m:
        return {"type": "other", "raw": raw_clean}

    msg = m.group(1)
    body = raw_clean[m.end():]
    parts = [msg] + [clean_token(x) for x in body.split("|")]
    while parts and parts[-1] == "":
        parts.pop()

    if msg == "DN":
        race = parts[1] if len(parts) > 1 else None
        heat = parts[2] if len(parts) > 2 else None
        return {"type": "new_run", "race": race, "heat": heat, "raw": raw_clean}

    if msg == "DA":
        race = parts[1] if len(parts) > 1 else None
        heat = parts[2] if len(parts) > 2 else None
        bibs = []
        b1 = extract_first_int(parts[3]) if len(parts) > 3 else None
        b2 = extract_first_int(parts[4]) if len(parts) > 4 else None
        if b1:
            bibs.append(b1)
        if b2:
            bibs.append(b2)
        return {"type": "setup", "race": race, "heat": heat, "bibs": bibs, "raw": raw_clean}

    if msg == "DS":
        race = parts[1] if len(parts) > 1 else None
        heat = parts[2] if len(parts) > 2 else None
        bibs = []
        b1 = extract_first_int(parts[3]) if len(parts) > 3 else None
        b2 = extract_first_int(parts[4]) if len(parts) > 4 else None
        if b1:
            bibs.append(b1)
        if b2:
            bibs.append(b2)

        start_time = None
        for t in parts[1:]:
            mm = TIME_HMS_RE.search(t)
            if mm:
                start_time = mm.group(0)
                break

        return {"type": "start", "race": race, "heat": heat, "bibs": bibs, "start_time": start_time, "raw": raw_clean}

    if msg in ("DI", "DF"):
        race = parts[1] if len(parts) > 1 else None
        heat = parts[2] if len(parts) > 2 else None
        split_no = extract_first_int(parts[3]) if len(parts) > 3 else None
        bib = extract_first_int(parts[4]) if len(parts) > 4 else None
        t = find_primary_time(parts[5:12])
        etype = "split" if msg == "DI" else "finish"
        return {"type": etype, "race": race, "heat": heat, "split": split_no, "bib": bib, "time": t, "raw": raw_clean}

    return {"type": "other", "raw": raw_clean}


class Athlete:
    def __init__(self, bib: str, name: str = "", country: str = ""):
        self.bib = bib
        self.name = name or ""
        self.country = (country or "").strip().upper()
        self.splits: Dict[str, float] = {}
        self.finish: Optional[float] = None
        self.status: str = ""
        self.pause_until: float = 0.0
        self.pause_value: Optional[float] = None

    def is_paused(self) -> bool:
        return time.monotonic() < self.pause_until and self.pause_value is not None




class Run:
    def __init__(self, race: str, heat: str, category: str = "DEFAULT"):
        self.category = clean_token(category or "") or "DEFAULT"
        self.race = race or "?"
        self.heat = heat or "?"
        self.key = f"{self.category}:{self.race}-{self.heat}"
        self.start_time: Optional[str] = None
        self.start_mono: Optional[float] = None
        self.athletes: Dict[str, Athlete] = {}
        self.prepared: bool = False
        self.active_bibs: List[str] = []
        self.bib_order: List[str] = []

    def ensure_athlete(self, bib: str, name: str = "", country: str = "") -> Optional[Athlete]:
        bib = clean_token(str(bib or ""))
        if not bib or bib == "0":
            return None
        if bib not in self.athletes:
            self.athletes[bib] = Athlete(bib, name=name or "", country=country or "")
            self.bib_order.append(bib)
        else:
            if name and not self.athletes[bib].name:
                self.athletes[bib].name = name
            if country and not self.athletes[bib].country:
                self.athletes[bib].country = (country or "").strip().upper()
        return self.athletes[bib]

    def finished_count(self) -> int:
        return sum(1 for a in self.athletes.values() if a.finish is not None)

    def total_count(self) -> int:
        return len(self.athletes)

    def split_ids(self) -> List[str]:
        ids = set()
        for a in self.athletes.values():
            ids.update(a.splits.keys())
        return sorted(ids, key=split_sort_key)



# 2) ЗАМЕНИТЬ class MeetModel целиком
class MeetModel:
    def __init__(self):
        self.runs: Dict[str, Run] = {}
        self.current_key: Optional[str] = None

        self.current_category: str = "DEFAULT"
        self.categories: set[str] = set(["DEFAULT"])

        self.bib_names: Dict[str, Dict[str, str]] = {"DEFAULT": {}}
        self.bib_country: Dict[str, Dict[str, str]] = {"DEFAULT": {}}

    def set_current_category(self, cat: str):
        cat = clean_token(cat or "") or "DEFAULT"
        self.current_category = cat
        self.categories.add(cat)
        self.bib_names.setdefault(cat, {})
        self.bib_country.setdefault(cat, {})

    def set_bib_meta(self, names, countries):
        self.bib_names = {}
        self.bib_country = {}
        self.categories = set()

        def _ensure_cat(cat: str):
            cat = clean_token(cat or "") or "DEFAULT"
            self.categories.add(cat)
            self.bib_names.setdefault(cat, {})
            self.bib_country.setdefault(cat, {})
            return cat

        if isinstance(names, dict) and names:
            any_val = next(iter(names.values()))
            if isinstance(any_val, dict):
                for cat, mp in names.items():
                    c = _ensure_cat(cat)
                    for k, v in (mp or {}).items():
                        self.bib_names[c][str(k)] = str(v or "").strip()
            else:
                c = _ensure_cat("DEFAULT")
                for k, v in (names or {}).items():
                    self.bib_names[c][str(k)] = str(v or "").strip()
        else:
            _ensure_cat("DEFAULT")

        if isinstance(countries, dict) and countries:
            any_val = next(iter(countries.values()))
            if isinstance(any_val, dict):
                for cat, mp in countries.items():
                    c = _ensure_cat(cat)
                    for k, v in (mp or {}).items():
                        self.bib_country[c][str(k)] = str(v or "").strip().upper()
            else:
                c = _ensure_cat("DEFAULT")
                for k, v in (countries or {}).items():
                    self.bib_country[c][str(k)] = str(v or "").strip().upper()

        if self.current_category not in self.categories:
            self.set_current_category(self.current_category)

    def set_bib_names(self, mapping: Dict[str, str]):
        self.set_bib_meta(mapping, {})

    def _name_for(self, cat: str, bib: str) -> str:
        cat = clean_token(cat or "") or "DEFAULT"
        return (self.bib_names.get(cat) or {}).get(str(bib), "")

    def _country_for(self, cat: str, bib: str) -> str:
        cat = clean_token(cat or "") or "DEFAULT"
        return (self.bib_country.get(cat) or {}).get(str(bib), "")

    def ensure_run(self, race, heat, category: Optional[str] = None) -> Run:
        cat = clean_token(category or self.current_category or "") or "DEFAULT"
        key = f"{cat}:{race or '?'}-{heat or '?'}"
        if key not in self.runs:
            self.runs[key] = Run(race, heat, cat)
        return self.runs[key]

    def _pick_run_for_bibs(self, race, heat, bibs: List[str], category: str) -> Optional[Run]:
        cur = self.runs.get(self.current_key) if self.current_key else None
        if cur and bibs:
            if set(cur.active_bibs) == set(bibs) or all(b in cur.athletes for b in bibs):
                return cur

        if bibs:
            for r in self.runs.values():
                if r.category != category:
                    continue
                if (race is None or r.race == (race or r.race)) and (heat is None or r.heat == (heat or r.heat)):
                    if set(r.active_bibs) == set(bibs) or all(b in r.athletes for b in bibs):
                        return r

        if bibs:
            for r in self.runs.values():
                if (race is None or r.race == (race or r.race)) and (heat is None or r.heat == (heat or r.heat)):
                    if set(r.active_bibs) == set(bibs) or all(b in r.athletes for b in bibs):
                        return r

        return None

    def apply(self, evt: Dict[str, Any]) -> Optional[str]:
        t = evt.get("type")
        race = evt.get("race")
        heat = evt.get("heat")
        cat = clean_token(evt.get("cat") or self.current_category or "") or "DEFAULT"
        self.set_current_category(cat)

        if t == "new_run":
            run = self.ensure_run(race, heat, cat)
            self.current_key = run.key
            return run.key

        if t == "setup":
            run = self.ensure_run(race, heat, cat)
            run.prepared = True

            bibs = evt.get("bibs") or []
            bibs = [b for b in bibs if b and str(b).strip() not in ("0", "")]

            if bibs:
                run.active_bibs = bibs[:2]
                for b in bibs:
                    a = run.ensure_athlete(b, self._name_for(cat, str(b)), self._country_for(cat, str(b)))
                    if a:
                        a.status = "готов"

                tail = [b for b in run.bib_order if b not in bibs]
                run.bib_order = list(bibs) + tail

            self.current_key = run.key
            return run.key

        if t == "start":
            bibs = evt.get("bibs") or []
            bibs = [b for b in bibs if b and str(b).strip() not in ("0", "")]

            run = self._pick_run_for_bibs(race, heat, bibs, cat) or self.ensure_run(race, heat, cat)
            run.start_time = evt.get("start_time") or run.start_time
            run.start_mono = time.monotonic()

            if bibs:
                run.active_bibs = bibs[:2]
                for b in bibs:
                    a = run.ensure_athlete(b, self._name_for(cat, str(b)), self._country_for(cat, str(b)))
                    if a:
                        a.status = "бежит"
                        a.pause_until = 0.0
                        a.pause_value = None

                tail = [b for b in run.bib_order if b not in bibs]
                run.bib_order = list(bibs) + tail

            self.current_key = run.key
            return run.key

        if t in ("split", "finish"):
            bib = evt.get("bib")

            run = None
            cur = self.runs.get(self.current_key) if self.current_key else None
            if cur and bib and (str(bib) in cur.athletes or str(bib) in cur.active_bibs):
                run = cur
            if run is None:
                run = self.ensure_run(race, heat, cat)

            ev_time = evt.get("time")
            if run.start_mono is None and ev_time is not None:
                run.start_mono = time.monotonic() - float(ev_time)

            a = run.ensure_athlete(bib, self._name_for(run.category, str(bib)), self._country_for(run.category, str(bib)))
            if not a:
                self.current_key = run.key
                return run.key

            now = time.monotonic()

            if t == "split":
                split_no = clean_token(str(evt.get("split") or ""))
                if split_no and ev_time is not None:
                    a.splits[split_no] = ev_time
                    a.status = f"отсечка {split_no}"
                else:
                    a.status = "отсечка"
                a.pause_value = ev_time
                a.pause_until = now + FREEZE_SEC
            else:
                a.finish = ev_time
                a.status = "финиш"
                a.pause_value = ev_time
                a.pause_until = float("inf")

            self.current_key = run.key
            return run.key

        return None

class TcpJsonlServer:
    def __init__(self, host: str, port: int, on_error=None):
        self.host = host
        self.port = port
        self.on_error = on_error
        self._stop = threading.Event()
        self._srv: Optional[socket.socket] = None
        self._clients: List[socket.socket] = []
        self._lock = threading.Lock()
        self._thr: Optional[threading.Thread] = None
        self._last_bytes: Optional[bytes] = None

    def start(self):
        if self._thr and self._thr.is_alive():
            return
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            s.bind((self.host, self.port))
            s.listen(16)
            s.settimeout(0.5)
            self._srv = s
        except Exception as e:
            if self.on_error:
                self.on_error(f"bind/listen failed: {e}")
            return

        self._stop.clear()
        self._thr = threading.Thread(target=self._accept_loop, daemon=True)
        self._thr.start()

    def stop(self):
        self._stop.set()
        try:
            if self._srv:
                self._srv.close()
        except Exception:
            pass
        with self._lock:
            for c in self._clients:
                try:
                    c.close()
                except Exception:
                    pass
            self._clients.clear()

    def _encode(self, obj: Dict[str, Any]) -> bytes:
        return (json.dumps(obj, ensure_ascii=False) + "\n").encode("utf-8", errors="ignore")

    def set_last(self, obj: Dict[str, Any]):
        try:
            self._last_bytes = self._encode(obj)
        except Exception:
            self._last_bytes = None

    def _accept_loop(self):
        s = self._srv
        if not s:
            return

        while not self._stop.is_set():
            try:
                c, _ = s.accept()
                c.setblocking(True)

                with self._lock:
                    self._clients.append(c)
                    lastb = self._last_bytes

                if lastb:
                    try:
                        c.sendall(lastb)
                    except Exception:
                        with self._lock:
                            try:
                                self._clients.remove(c)
                            except Exception:
                                pass
                        try:
                            c.close()
                        except Exception:
                            pass

            except socket.timeout:
                continue
            except Exception as e:
                if self.on_error:
                    self.on_error(f"accept loop error: {e}")
                break

    def broadcast(self, obj: Dict[str, Any]):
        self.set_last(obj)
        data = self._last_bytes
        if not data:
            return

        dead = []
        with self._lock:
            for c in self._clients:
                try:
                    c.sendall(data)
                except Exception:
                    dead.append(c)

            for c in dead:
                try:
                    c.close()
                except Exception:
                    pass
                try:
                    self._clients.remove(c)
                except Exception:
                    pass


class TcpStateServer:
    def __init__(self, host: str, port: int, on_error=None):
        self.host = host
        self.port = port
        self.on_error = on_error
        self._stop = threading.Event()
        self._srv: Optional[socket.socket] = None
        self._clients: List[socket.socket] = []
        self._lock = threading.Lock()
        self._thr: Optional[threading.Thread] = None
        self._last_state_bytes: Optional[bytes] = None

    def start(self):
        if self._thr and self._thr.is_alive():
            return
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            s.bind((self.host, self.port))
            s.listen(16)
            s.settimeout(0.5)
            self._srv = s
        except Exception as e:
            if self.on_error:
                self.on_error(f"bind/listen failed: {e}")
            return

        self._stop.clear()
        self._thr = threading.Thread(target=self._accept_loop, daemon=True)
        self._thr.start()

    def stop(self):
        self._stop.set()
        try:
            if self._srv:
                self._srv.close()
        except Exception:
            pass
        with self._lock:
            for c in self._clients:
                try:
                    c.close()
                except Exception:
                    pass
            self._clients.clear()

    def _encode_state(self, state: Dict[str, Any]) -> bytes:
        line = json.dumps({"type": "state", "state": state}, ensure_ascii=False) + "\n"
        return line.encode("utf-8", errors="ignore")

    def set_last_state(self, state: Dict[str, Any]):
        try:
            self._last_state_bytes = self._encode_state(state)
        except Exception:
            self._last_state_bytes = None

    def _accept_loop(self):
        s = self._srv
        if not s:
            return

        while not self._stop.is_set():
            try:
                c, _ = s.accept()
                c.setblocking(True)

                try:
                    hello = (json.dumps({"type": "hello", "v": 1}, ensure_ascii=False) + "\n").encode("utf-8")
                    c.sendall(hello)
                except Exception:
                    try:
                        c.close()
                    except Exception:
                        pass
                    continue

                with self._lock:
                    self._clients.append(c)
                    lastb = self._last_state_bytes

                if lastb:
                    try:
                        c.sendall(lastb)
                    except Exception:
                        with self._lock:
                            try:
                                self._clients.remove(c)
                            except Exception:
                                pass
                        try:
                            c.close()
                        except Exception:
                            pass

            except socket.timeout:
                continue
            except Exception as e:
                if self.on_error:
                    self.on_error(f"accept loop error: {e}")
                break

    def broadcast_state(self, state: Dict[str, Any]):
        self.set_last_state(state)
        data = self._last_state_bytes
        if not data:
            return

        dead = []
        with self._lock:
            for c in self._clients:
                try:
                    c.sendall(data)
                except Exception:
                    dead.append(c)
            for c in dead:
                try:
                    c.close()
                except Exception:
                    pass
                try:
                    self._clients.remove(c)
                except Exception:
                    pass


class ReaderThread(threading.Thread):
    def __init__(self, q, stop_evt, port=None, baud=9600, replay_path=None):
        super().__init__(daemon=True)
        self.q = q
        self.stop_evt = stop_evt
        self.port = port
        self.baud = baud
        self.replay_path = replay_path
        self.ser = None

    def _emit_evt(self, evt):
        self.q.put({"kind": "evt", "data": evt})

    def run(self):
        if self.replay_path:
            self._run_replay()
        else:
            self._run_serial()

    def _drain_buf(self, buf: str):
        while True:
            msgs, buf2 = split_stream(buf)
            if not msgs:
                return buf
            for raw in msgs:
                evt = parse_message(raw)
                if evt:
                    self._emit_evt(evt)
            buf = buf2

    def _run_replay(self):
        try:
            with open(self.replay_path, "r", encoding="utf-8", errors="ignore") as f:
                buf = ""
                while not self.stop_evt.is_set():
                    chunk = f.read(512)
                    if not chunk:
                        break
                    buf += chunk
                    buf = self._drain_buf(buf)
                    time.sleep(0.002)

                tail = buf.strip()
                if tail:
                    evt = parse_message(tail)
                    if evt:
                        self._emit_evt(evt)

        except Exception as e:
            self.q.put({"kind": "err", "data": str(e)})

    def _run_serial(self):
        if serial is None:
            self.q.put({"kind": "err", "data": "pyserial не установлен"})
            return

        try:
            self.ser = serial.Serial(
                self.port,
                self.baud,
                timeout=0.2,
                bytesize=serial.EIGHTBITS,
                parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,
                xonxoff=False,
                rtscts=False,
                dsrdtr=False,
            )
        except Exception as e:
            self.q.put({"kind": "err", "data": f"Не открыл порт: {e}"})
            return

        buf = ""
        last_byte_ts = time.time()

        try:
            while not self.stop_evt.is_set():
                try:
                    n = self.ser.in_waiting
                except Exception:
                    n = 0

                data = self.ser.read(n if n else 1)

                if data:
                    last_byte_ts = time.time()
                    s = data.decode("ascii", errors="ignore")
                    buf += s
                    buf = self._drain_buf(buf)
                else:
                    if buf and (time.time() - last_byte_ts) > 0.35:
                        tail = buf.strip()
                        if tail and MSG_RE.search(tail):
                            evt = parse_message(tail)
                            if evt:
                                self._emit_evt(evt)
                            buf = ""

        except Exception as e:
            self.q.put({"kind": "err", "data": str(e)})
        finally:
            try:
                self.ser.close()
            except Exception:
                pass


class App(tk.Tk):
    def __init__(self, listen_host: str, listen_port: int):
        super().__init__()
        self.title("SwissTiming Quantum Viewer (NET)")
        self.geometry("1360x860")
        self.minsize(1100, 720)

        self.q = queue.Queue()
        self.stop_evt = threading.Event()
        self.reader = None

        self.model = MeetModel()
        self.run_items: Dict[str, str] = {}
        self.selected_run_key: Optional[str] = None

        self.port_var = tk.StringVar()
        self.baud_var = tk.StringVar(value="9600")
        self.status_var = tk.StringVar(value="Отключено")
        self.status_color = "#ff5c5c"
        self.category_var = tk.StringVar(value="DEFAULT")
        self.category_var.trace_add("write", lambda *_: self.model.set_current_category(self.category_var.get()))

        self.ath_split_ids: List[str] = []

        self.live_run_var = tk.StringVar(value="Заезд: —")
        self.live_bib1_var = tk.StringVar(value="")
        self.live_time1_var = tk.StringVar(value="")
        self.live_bib2_var = tk.StringVar(value="")
        self.live_time2_var = tk.StringVar(value="")

        self._obs_lock = threading.Lock()
        self._obs_payload = {"ts": 0, "run": "", "left": {}, "right": {}}

        # overlay.html лежит рядом с программой/скриптом
        overlay_path = resource_path("overlay_test.html")
        flags_dir = resource_path("flags")
        #overlay_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "overlay.html")
        self.overlay_http = OverlayHttp(self, host="0.0.0.0", port=8099, overlay_html_path=overlay_path,flags_dir=flags_dir)
        self.overlay_http.start()

        self._colors = {
            "bg": "#0f1117",
            "panel": "#151a23",
            "panel2": "#111621",
            "fg": "#e6e6e6",
            "muted": "#a8b0bf",
            "line": "#242b3a",
            "accent": "#3aa0ff",
            "accent2": "#7ee787",
            "danger": "#ff5c5c",
            "select": "#243044",
            "head": "#1b2230",
            "odd": "#121826",
            "even": "#0f1522",
        }
        #HTTP Server
        self.net = TcpStateServer(listen_host, listen_port, on_error=self._net_error)
        self.net.start()
        self.net.set_last_state(self._model_to_state())
        #TCP JSON Server
        self.live_tcp = TcpJsonlServer(listen_host, 8098, on_error=self._net_error)
        self.live_tcp.start()
        self.live_tcp.set_last({"ts": 0, "run": "", "left": {}, "right": {}})



        self._setup_style()
        self._build_ui()

        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self.after(50, self._pump)

        self.obs_out = "C:\\Users\\Lokosphinx\\PycharmProjects\\SwissData\\obs_live.txt"
        self.obs_json = "C:\\Users\\Lokosphinx\\PycharmProjects\\SwissData\\obs_state.json"

    def demo_start(self):
        # создаём один заезд и двух гонщиков
        run = self.model.ensure_run("1", "1")
        run.prepared = True
        run.active_bibs = ["97", "852"]
        run.start_time = time.strftime("%H:%M:%S.000")
        run.start_mono = time.monotonic()
        self.model.current_key = run.key

        a1 = run.ensure_athlete("97", self.model.bib_names.get("97", "Бурьянов Дмитрий"))
        a2 = run.ensure_athlete("852", self.model.bib_names.get("852", "Васильев Иван"))
        if a1: a1.status = "бежит"
        if a2: a2.status = "бежит"

        # раз в 3 секунды делаем “отсечку” то одному, то другому (чтобы проверить freeze)
        self._demo_flip = False

        def _demo_tick():
            if not run.start_mono:
                return
            t = max(0.0, time.monotonic() - run.start_mono)
            self._demo_flip = not self._demo_flip
            if self._demo_flip and a1:
                a1.splits[str(len(a1.splits) + 1)] = t
                a1.pause_value = t
                a1.pause_until = time.monotonic() + FREEZE_SEC
            elif (not self._demo_flip) and a2:
                a2.splits[str(len(a2.splits) + 1)] = t
                a2.pause_value = t
                a2.pause_until = time.monotonic() + FREEZE_SEC

            self.net.broadcast_state(self._model_to_state())
            self.after(3000, _demo_tick)

        self.net.broadcast_state(self._model_to_state())
        self.after(3000, _demo_tick)

    def _net_error(self, msg: str):
        try:
            self._append_log("NET ERROR: " + msg)
        except Exception:
            pass
        try:
            self._set_status("NET ERROR", False)
        except Exception:
            pass

    def _on_close(self):
        try:
            self.live_tcp.stop()
        except Exception:
            pass
        try:
            with self._obs_lock:
                self._obs_payload = {"ts": 0, "run": "", "left": {}, "right": {}}
            self.overlay_http.stop()

        except Exception:
            pass
        try:
            self.disconnect()
        except Exception:
            pass
        try:
            self.net.stop()
        except Exception:
            pass
        self.destroy()

    def _model_to_state(self) -> Dict[str, Any]:
        runs_out: Dict[str, Any] = {}
        for k, run in self.model.runs.items():
            ath = {}
            for bib, a in run.athletes.items():
                ath[bib] = {
                    "bib": bib,
                    "name": a.name,
                    "splits": dict(a.splits),
                    "finish": a.finish,
                    "status": a.status,
                }
            runs_out[k] = {
                "key": run.key,
                "category": run.category,
                "race": run.race,
                "heat": run.heat,
                "start_time": run.start_time,
                "active_bibs": list(run.active_bibs),
                "bib_order": list(run.bib_order),
                "athletes": ath,
            }
        return {
            "current_key": self.model.current_key,
            "current_category": self.model.current_category,
            "runs": runs_out,
            "ts": time.time(),
        }

    def _setup_style(self):
        c = self._colors
        self.configure(bg=c["bg"])
        style = ttk.Style()
        style.theme_use("clam")

        try:
            import tkinter.font as tkfont
            f = tkfont.nametofont("TkDefaultFont")
            f.configure(family="Segoe UI", size=12)
            f2 = tkfont.nametofont("TkTextFont")
            f2.configure(family="Segoe UI", size=12)
        except Exception:
            pass

        style.configure(".", background=c["bg"], foreground=c["fg"])
        style.configure("TFrame", background=c["bg"])
        style.configure("Card.TFrame", background=c["panel"])
        style.configure("TimerCard.TFrame", background=c["panel2"])
        style.configure("TLabel", background=c["bg"], foreground=c["fg"])
        style.configure("Muted.TLabel", background=c["bg"], foreground=c["muted"])
        style.configure("Title.TLabel", background=c["bg"], foreground=c["fg"], font=("Segoe UI", 16, "bold"))
        style.configure("H2.TLabel", background=c["bg"], foreground=c["fg"], font=("Segoe UI", 13, "bold"))

        style.configure("TimerTitle.TLabel", background=c["panel2"], foreground=c["muted"], font=("Segoe UI", 12, "bold"))
        style.configure("TimerRun.TLabel", background=c["panel2"], foreground=c["muted"], font=("Segoe UI", 12))
        style.configure("AthBib.TLabel", background=c["panel2"], foreground=c["fg"], font=("Segoe UI", 18, "bold"))
        style.configure("AthTime.TLabel", background=c["panel2"], foreground=c["fg"], font=("Segoe UI", 34, "bold"))

        style.configure("TButton", background=c["panel"], foreground=c["fg"], borderwidth=0, padding=(14, 10))
        style.map("TButton", background=[("active", c["head"]), ("pressed", c["select"])])

        style.configure("Accent.TButton", background=c["accent"], foreground="#0b0d12", padding=(14, 10))
        style.map("Accent.TButton", background=[("active", "#5bb3ff"), ("pressed", "#2f8fe6")])

        style.configure("TCombobox",
                        padding=(10, 8),
                        fieldbackground=c["panel2"],
                        background=c["panel2"],
                        foreground=c["fg"],
                        arrowcolor=c["fg"])

        style.configure("TNotebook", background=c["bg"], borderwidth=0)
        style.configure("TNotebook.Tab", background=c["panel"], foreground=c["muted"], padding=(14, 10), borderwidth=0)
        style.map("TNotebook.Tab", background=[("selected", c["head"])], foreground=[("selected", c["fg"])])

        style.configure("Treeview",
                        background=c["panel2"],
                        fieldbackground=c["panel2"],
                        foreground=c["fg"],
                        rowheight=38,
                        borderwidth=0)
        style.map("Treeview",
                  background=[("selected", c["select"])],
                  foreground=[("selected", c["fg"])])

        style.configure("Treeview.Heading",
                        background=c["head"],
                        foreground=c["fg"],
                        relief="flat",
                        padding=(10, 10),
                        borderwidth=0,
                        font=("Segoe UI", 12, "bold"))

    def _list_ports(self):
        if list_ports is None:
            return []
        return [p.device for p in list_ports.comports()]

    def _build_ui(self):
        c = self._colors

        root = ttk.Frame(self, style="TFrame")
        root.pack(fill="both", expand=True)

        top = ttk.Frame(root, style="TFrame")
        top.pack(fill="x", padx=14, pady=(14, 10))
        ttk.Label(top, text="SwissTiming Quantum Viewer", style="Title.TLabel").pack(anchor="w")

        bar = ttk.Frame(root, style="Card.TFrame")
        bar.pack(fill="x", padx=14, pady=(0, 12))

        bar_inner = ttk.Frame(bar, style="Card.TFrame")
        bar_inner.pack(fill="x", padx=12, pady=12)

        ttk.Label(bar_inner, text="COM", style="Muted.TLabel").pack(side="left")
        self.port_cb = ttk.Combobox(bar_inner, textvariable=self.port_var, width=12,
                                    values=self._list_ports(), state="readonly")
        self.port_cb.pack(side="left", padx=(8, 14))

        ttk.Label(bar_inner, text="Baud", style="Muted.TLabel").pack(side="left")
        self.baud_cb = ttk.Combobox(bar_inner, textvariable=self.baud_var, width=10,
                                    values=["9600", "19200", "38400", "57600", "115200"], state="readonly")
        self.baud_cb.pack(side="left", padx=(8, 14))

        ttk.Button(bar_inner, text="Обновить", command=self._refresh_ports).pack(side="left", padx=(0, 10))
        ttk.Button(bar_inner, text="Подключить", style="Accent.TButton", command=self.connect).pack(side="left", padx=(0, 10))
        ttk.Button(bar_inner, text="Отключить", command=self.disconnect).pack(side="left", padx=(0, 10))
        ttk.Button(bar_inner, text="Файл", command=self.replay_file).pack(side="left", padx=(0, 10))
        ttk.Button(bar_inner, text="Excel", command=self.load_excel).pack(side="left", padx=(0, 10))
        ttk.Label(bar_inner, text="Категория:", style="Muted.TLabel").pack(side="left")
        self.cat_cb = ttk.Combobox(
            bar_inner,
            textvariable=self.category_var,
            width=16,
            values=sorted(list(self.model.categories)),
            state="normal"
        )
        self.cat_cb.pack(side="left", padx=(8, 16))
        #ttk.Button(bar_inner, text="CSV", command=self.export_csv).pack(side="left", padx=(0, 10))
        #ttk.Button(bar_inner, text="DEMO", command=self.demo_start).pack(side="left", padx=(0, 10))

        status_wrap = ttk.Frame(bar_inner, style="Card.TFrame")
        status_wrap.pack(side="right")

        self.status_dot = tk.Canvas(status_wrap, width=14, height=14, bg=c["panel"], highlightthickness=0)
        self.status_dot.pack(side="left", padx=(0, 8))
        self._dot_id = self.status_dot.create_oval(2, 2, 12, 12, fill=self.status_color, outline=self.status_color)
        ttk.Label(status_wrap, textvariable=self.status_var, style="H2.TLabel").pack(side="left")

        self.nb = ttk.Notebook(root)
        self.nb.pack(fill="both", expand=True, padx=14, pady=(0, 14))

        tab_results = ttk.Frame(self.nb, style="TFrame")
        tab_log = ttk.Frame(self.nb, style="TFrame")
        self.nb.add(tab_results, text="Результаты")
        self.nb.add(tab_log, text="Сырые данные")

        pan = ttk.Panedwindow(tab_results, orient="horizontal")
        pan.pack(fill="both", expand=True)

        left = ttk.Frame(pan, style="Card.TFrame")
        right = ttk.Frame(pan, style="Card.TFrame")
        pan.add(left, weight=1)
        pan.add(right, weight=4)

        left_pad = ttk.Frame(left, style="Card.TFrame")
        left_pad.pack(fill="both", expand=True, padx=12, pady=12)
        ttk.Label(left_pad, text="Заезды", style="H2.TLabel").pack(anchor="w", pady=(0, 8))

        self.run_filter_var = tk.StringVar()
        filter_row = ttk.Frame(left_pad, style="Card.TFrame")
        filter_row.pack(fill="x", pady=(0, 10))
        ttk.Label(filter_row, text="Фильтр", style="Muted.TLabel").pack(side="left")
        self.run_filter_entry = tk.Entry(filter_row, textvariable=self.run_filter_var,
                                         bg=c["panel2"], fg=c["fg"], insertbackground=c["fg"],
                                         relief="flat", highlightthickness=1,
                                         highlightbackground=c["line"], highlightcolor=c["accent"])
        self.run_filter_entry.pack(side="left", fill="x", expand=True, padx=(8, 0), ipady=6)
        self.run_filter_var.trace_add("write", lambda *_: self._render_runs())

        self.runs_tv = ttk.Treeview(left_pad, columns=("run", "start", "ath", "fin"), show="headings", height=18)
        for col, text, w in [
            ("run", "Заезд", 90),
            ("start", "Старт", 140),
            ("ath", "Участн.", 90),
            ("fin", "Финиш", 80),
        ]:
            self.runs_tv.heading(col, text=text)
            self.runs_tv.column(col, width=w, anchor="center")

        self.runs_tv.tag_configure("odd", background=c["odd"])
        self.runs_tv.tag_configure("even", background=c["even"])

        runs_wrap = ttk.Frame(left_pad, style="Card.TFrame")
        runs_wrap.pack(fill="both", expand=True)
        runs_vsb = ttk.Scrollbar(runs_wrap, orient="vertical", command=self.runs_tv.yview)
        self.runs_tv.configure(yscrollcommand=runs_vsb.set)
        self.runs_tv.pack(side="left", fill="both", expand=True)
        runs_vsb.pack(side="right", fill="y")
        self.runs_tv.bind("<<TreeviewSelect>>", self._on_run_select)

        right_pad = ttk.Frame(right, style="Card.TFrame")
        right_pad.pack(fill="both", expand=True, padx=12, pady=12)

        head_row = ttk.Frame(right_pad, style="Card.TFrame")
        head_row.pack(fill="x")
        ttk.Label(head_row, text="Участники", style="H2.TLabel").pack(side="left")

        self.run_info_var = tk.StringVar(value="—")
        ttk.Label(head_row, textvariable=self.run_info_var, style="Muted.TLabel").pack(side="right")

        timer_card = ttk.Frame(right_pad, style="TimerCard.TFrame")
        timer_card.pack(fill="x", pady=(10, 10))
        timer_inner = ttk.Frame(timer_card, style="TimerCard.TFrame")
        timer_inner.pack(fill="x", padx=12, pady=12)

        ttk.Label(timer_inner, text="ТЕКУЩЕЕ ВРЕМЯ (2 ГОНЩИКА)", style="TimerTitle.TLabel").pack(anchor="w")
        ttk.Label(timer_inner, textvariable=self.live_run_var, style="TimerRun.TLabel").pack(anchor="w", pady=(2, 10))

        row = ttk.Frame(timer_inner, style="TimerCard.TFrame")
        row.pack(fill="x")
        left_box = ttk.Frame(row, style="TimerCard.TFrame")
        right_box = ttk.Frame(row, style="TimerCard.TFrame")
        left_box.grid(row=0, column=0, sticky="ew", padx=(0, 14))
        right_box.grid(row=0, column=1, sticky="ew")
        row.columnconfigure(0, weight=1)
        row.columnconfigure(1, weight=1)

        ttk.Label(left_box, textvariable=self.live_bib1_var, style="AthBib.TLabel").pack(anchor="w")
        ttk.Label(left_box, textvariable=self.live_time1_var, style="AthTime.TLabel").pack(anchor="w", pady=(2, 0))
        ttk.Label(right_box, textvariable=self.live_bib2_var, style="AthBib.TLabel").pack(anchor="w")
        ttk.Label(right_box, textvariable=self.live_time2_var, style="AthTime.TLabel").pack(anchor="w", pady=(2, 0))

        self.ath_container = ttk.Frame(right_pad, style="Card.TFrame")
        self.ath_container.pack(fill="both", expand=True)
        self._rebuild_ath_tree(split_ids=[])

        log_card = ttk.Frame(tab_log, style="Card.TFrame")
        log_card.pack(fill="both", expand=True, padx=12, pady=12)
        log_pad = ttk.Frame(log_card, style="Card.TFrame")
        log_pad.pack(fill="both", expand=True, padx=12, pady=12)

        ttk.Label(log_pad, text="Сырые данные (последние 500 строк)", style="H2.TLabel").pack(anchor="w", pady=(0, 10))
        self.log = tk.Text(log_pad, height=10, wrap="none",
                           bg=c["panel2"], fg=c["fg"], insertbackground=c["fg"],
                           relief="flat", highlightthickness=1,
                           highlightbackground=c["line"], highlightcolor=c["accent"])
        self.log.pack(fill="both", expand=True)

    def _set_status(self, text: str, ok: bool):
        self.status_var.set(text)
        self.status_color = self._colors["accent2"] if ok else self._colors["danger"]
        try:
            self.status_dot.itemconfigure(self._dot_id, fill=self.status_color, outline=self.status_color)
        except Exception:
            pass

    def _refresh_ports(self):
        self.port_cb["values"] = self._list_ports()

    def connect(self):
        if self.reader and self.reader.is_alive():
            return
        port = self.port_var.get().strip()
        if not port:
            messagebox.showerror("Ошибка", "Выбери COM-порт")
            return
        try:
            baud = int(self.baud_var.get().strip())
        except Exception:
            messagebox.showerror("Ошибка", "Неверный baud")
            return
        self.stop_evt.clear()
        self.reader = ReaderThread(self.q, self.stop_evt, port=port, baud=baud, replay_path=None)
        self.reader.start()
        self._set_status(f"Подключено {port}@{baud}", True)

    def disconnect(self):
        self.stop_evt.set()
        self._set_status("Отключено", False)

    def replay_file(self):
        path = filedialog.askopenfilename(title="Выбери лог", filetypes=[("Text", "*.txt"), ("All", "*.*")])
        if not path:
            return
        self.disconnect()
        self.model = MeetModel()
        self.run_items.clear()
        self.selected_run_key = None
        self.run_filter_var.set("")
        self._render_runs()
        self._rebuild_ath_tree(split_ids=[])
        self._render_athletes(None)
        self.stop_evt.clear()
        self.reader = ReaderThread(self.q, self.stop_evt, replay_path=path)
        self.reader.start()
        self._set_status("Режим файла", True)
        self.net.broadcast_state(self._model_to_state())

    def load_excel(self):
        if load_workbook is None:
            messagebox.showerror("Ошибка", "Нужен openpyxl (pip install openpyxl)")
            return
        path = filedialog.askopenfilename(title="Выбери Excel", filetypes=[("Excel", "*.xlsx"), ("All", "*.*")])
        if not path:
            return
        try:
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            mapping_name: Dict[str, Dict[str, str]] = {}
            mapping_country: Dict[str, Dict[str, str]] = {}

            for r in ws.iter_rows(min_row=1, max_col=4, values_only=True):
                bib_val = r[0]
                name_val = r[1] if len(r) > 1 else None
                country_val = r[2] if len(r) > 2 else None
                cat_val = r[3] if len(r) > 3 else None
                if bib_val is None:
                    continue
                cat_s = clean_token(cat_val) or "DEFAULT"
                bib_s = None
                if isinstance(bib_val, (int, float)):
                    try:
                        bib_i = int(bib_val)
                        if bib_i != 0:
                            bib_s = str(bib_i)
                    except Exception:
                        bib_s = None
                else:
                    m = re.search(r"(\d+)", str(bib_val))
                    if m:
                        try:
                            bib_i = int(m.group(1))
                            if bib_i != 0:
                                bib_s = str(bib_i)
                        except Exception:
                            bib_s = None

                if not bib_s:
                    continue

                name_s = str(name_val).strip() if name_val is not None else ""
                cc_s = str(country_val).strip().upper() if country_val is not None else ""

                mapping_name.setdefault(cat_s, {})[bib_s] = name_s
                if cc_s:
                    mapping_country.setdefault(cat_s, {})[bib_s] = cc_s

            self.model.set_bib_meta(mapping_name, mapping_country)

            for run in self.model.runs.values():
                for bib, a in run.athletes.items():
                    nm = mapping_name.get(run.category, {}).get(str(bib), "")
                    cc = mapping_country.get(run.category, {}).get(str(bib), "")
                    if nm:
                        a.name = nm
                    if cc:
                        a.country = cc
            cats = sorted(list(self.model.categories))
            self.cat_cb["values"] = cats
            if cats and self.category_var.get().strip() not in cats:
                self.category_var.set(cats[0])
            if self.selected_run_key:
                self._render_athletes(self.selected_run_key)

            self.net.broadcast_state(self._model_to_state())

        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def export_csv(self):
        path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV", "*.csv")])
        if not path:
            return
        try:
            all_splits = set()
            for run in self.model.runs.values():
                all_splits.update(run.split_ids())
            all_splits_sorted = sorted(all_splits, key=split_sort_key)

            cols = ["run", "bib", "name"] + [f"S{sid}" for sid in all_splits_sorted] + ["finish", "status"]

            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(cols)
                for run_key in list(self.model.runs.keys()):
                    run = self.model.runs[run_key]
                    order = [b for b in run.bib_order if b in run.athletes]
                    for bib in order:
                        a = run.athletes[bib]
                        row = [run.key, a.bib, a.name]
                        for sid in all_splits_sorted:
                            row.append(fmt_time(a.splits.get(str(sid))))
                        row.append(fmt_time(a.finish))
                        row.append(a.status)
                        w.writerow(row)
        except Exception as e:
            messagebox.showerror("Ошибка", str(e))

    def _append_log(self, s):
        s = strip_ctrl(s).rstrip()
        if not s:
            return
        self.log.insert("end", s + "\n")
        lines = int(self.log.index("end-1c").split(".")[0])
        if lines > 500:
            self.log.delete("1.0", f"{lines-500}.0")
        self.log.see("end")

    def _on_run_select(self, _evt=None):
        sel = self.runs_tv.selection()
        if not sel:
            self.selected_run_key = None
            self._render_athletes(None)
            return
        iid = sel[0]
        key = None
        for k, v in self.run_items.items():
            if v == iid:
                key = k
                break
        self.selected_run_key = key
        self._render_athletes(key)

    def _render_runs(self):
        for iid in self.runs_tv.get_children():
            self.runs_tv.delete(iid)
        self.run_items.clear()

        flt = self.run_filter_var.get().strip().lower()
        keys = list(self.model.runs.keys())
        idx = 0

        for run_key in keys:
            run = self.model.runs[run_key]
            if flt:
                hay = f"{run.key} {run.start_time or ''} {run.total_count()} {run.finished_count()}".lower()
                if flt not in hay:
                    continue

            tag = "even" if (idx % 2 == 0) else "odd"
            iid = self.runs_tv.insert(
                "", "end",
                values=(run.key, run.start_time or "", run.total_count(), run.finished_count()),
                tags=(tag,)
            )
            self.run_items[run_key] = iid
            idx += 1

    def _render_or_update_run_row(self, run_key: str):
        run = self.model.runs.get(run_key)
        if not run:
            return
        if run_key not in self.run_items:
            self._render_runs()
            return
        iid = self.run_items[run_key]
        self.runs_tv.item(iid, values=(run.key, run.start_time or "", run.total_count(), run.finished_count()))

    def _rebuild_ath_tree(self, split_ids: List[str]):
        self.ath_split_ids = list(split_ids)
        c = self._colors

        for child in self.ath_container.winfo_children():
            child.destroy()

        cols = ["bib", "name"] + [safe_col_id(sid) for sid in split_ids] + ["finish", "status"]

        self.ath_tv = ttk.Treeview(self.ath_container, columns=cols, show="headings")
        vsb = ttk.Scrollbar(self.ath_container, orient="vertical", command=self.ath_tv.yview)
        hsb = ttk.Scrollbar(self.ath_container, orient="horizontal", command=self.ath_tv.xview)
        self.ath_tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.ath_tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        self.ath_container.rowconfigure(0, weight=1)
        self.ath_container.columnconfigure(0, weight=1)

        self.ath_tv.heading("bib", text="№")
        self.ath_tv.column("bib", width=90, anchor="center")
        self.ath_tv.heading("name", text="Имя")
        self.ath_tv.column("name", width=240, anchor="w")

        for sid in split_ids:
            cid = safe_col_id(sid)
            self.ath_tv.heading(cid, text=f"S{sid}")
            self.ath_tv.column(cid, width=130, anchor="center")

        self.ath_tv.heading("finish", text="Финиш")
        self.ath_tv.column("finish", width=150, anchor="center")

        self.ath_tv.heading("status", text="Статус")
        self.ath_tv.column("status", width=180, anchor="center")

        self.ath_tv.tag_configure("odd", background=c["odd"])
        self.ath_tv.tag_configure("even", background=c["even"])

    def _render_athletes(self, run_key: Optional[str]):
        run = self.model.runs.get(run_key) if run_key else None

        if run:
            desired = run.split_ids()
            if desired != self.ath_split_ids:
                self._rebuild_ath_tree(split_ids=desired)
            self.run_info_var.set(f"{run.key}   старт: {run.start_time or '—'}   участников: {run.total_count()}   финиш: {run.finished_count()}")
        else:
            if self.ath_split_ids:
                self._rebuild_ath_tree(split_ids=[])
            self.run_info_var.set("—")

        for iid in self.ath_tv.get_children():
            self.ath_tv.delete(iid)

        if not run:
            return

        split_ids = self.ath_split_ids
        idx = 0
        order = [b for b in run.bib_order if b in run.athletes]
        for bib in order:
            a = run.athletes[bib]
            values = [a.bib, a.name]
            for sid in split_ids:
                values.append(fmt_time(a.splits.get(str(sid))))
            values.append(fmt_time(a.finish))
            values.append(a.status)
            tag = "even" if (idx % 2 == 0) else "odd"
            self.ath_tv.insert("", "end", values=values, tags=(tag,))
            idx += 1

    def _athlete_display_live(self, run: Run, a: Optional[Athlete]) -> Optional[float]:
        if not run or a is None:
            return None
        if run.start_mono is None:
            return None
        if a.finish is not None:
            return a.finish
        if a.is_paused():
            return a.pause_value
        return max(0.0, time.monotonic() - run.start_mono)

    def _tick_live_panel(self):
        key = self.selected_run_key or self.model.current_key
        run = self.model.runs.get(key) if key else None

        if not run:
            self.live_run_var.set("Заезд: —")
            self.live_bib1_var.set("")
            self.live_bib2_var.set("")
            self.live_time1_var.set("")
            self.live_time2_var.set("")
            with self._obs_lock:
                self._obs_payload = {"ts": 0, "run": "", "left": {}, "right": {}}
            try:
                _atomic_write_text(self.obs_json, json.dumps(self._obs_payload, ensure_ascii=False))
            except Exception:
                pass
            try:
                self.live_tcp.broadcast({"ts": 0, "run": "", "left": {}, "right": {}})
            except Exception:
                pass
            return

        self.live_run_var.set(f"Заезд: {run.key}")

        b1 = run.active_bibs[0] if len(run.active_bibs) > 0 else ""
        b2 = run.active_bibs[1] if len(run.active_bibs) > 1 else ""

        a1 = run.athletes.get(b1) if b1 else None
        a2 = run.athletes.get(b2) if b2 else None

        t1 = self._athlete_display_live(run, a1) if b1 else None
        t2 = self._athlete_display_live(run, a2) if b2 else None

        self.live_bib1_var.set(b1 or "")
        self.live_bib2_var.set(b2 or "")
        self.live_time1_var.set(fmt_live(t1) if b1 else "")
        self.live_time2_var.set(fmt_live(t2) if b2 else "")

        start_epoch = None
        if run.start_mono is not None:
            start_epoch = time.time() - (time.monotonic() - run.start_mono)

        def pack_live(a: Optional[Athlete], bib: str) -> dict:
            if not a or not bib:
                return {}

            country = (a.country or "").strip()
            if a.pause_until == 0 or a.pause_until is None:
                paused_until_epoch = None
            elif a.pause_until == float("inf"):
                paused_until_epoch = 10 ** 18
            else:
                paused_until_epoch = _mono_to_epoch(a.pause_until)

            return {
                "bib": bib,
                "name": (a.name or "").strip(),
                "country": country,
                "flag": flag_for_excel_value(country),
                "start_epoch": start_epoch,
                "paused_until_epoch": paused_until_epoch,
                "paused_value": a.pause_value,
                "finish": a.finish,
            }

        payload = {
            "ts": time.time(),
            "run": run.key,
            "left": pack_live(a1, b1),
            "right": pack_live(a2, b2),
        }
        try:
            self.live_tcp.broadcast(payload)
        except Exception:
            pass


        with self._obs_lock:
            self._obs_payload = payload

        try:
            _atomic_write_text(self.obs_json, json.dumps(payload, ensure_ascii=False))
        except Exception:
            pass

    def _apply_evt(self, evt):
        evt = dict(evt or {})
        if not evt.get("cat"):
            evt["cat"] = clean_token(self.category_var.get()) or "DEFAULT"
        run_key = self.model.apply(evt)
        self._append_log(evt.get("raw", ""))

        if run_key:
            if not self.run_items:
                self._render_runs()
            else:
                self._render_or_update_run_row(run_key)

            if self.selected_run_key == run_key:
                self._render_athletes(run_key)

            if self.selected_run_key is None:
                self.selected_run_key = run_key
                self._render_runs()
                iid = self.run_items.get(run_key)
                if iid:
                    self.runs_tv.selection_set(iid)
                    self.runs_tv.see(iid)
                self._render_athletes(run_key)

        self.net.broadcast_state(self._model_to_state())

    def _pump(self):
        try:
            while True:
                item = self.q.get_nowait()
                if item["kind"] == "evt":
                    self._apply_evt(item["data"])
                elif item["kind"] == "err":
                    self._append_log("ERROR: " + item["data"])
                    self._set_status("Ошибка", False)
        except queue.Empty:
            pass

        self._tick_live_panel()
        self.after(50, self._pump)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--listen", default="0.0.0.0")
    ap.add_argument("--port", type=int, default=9876)
    ap.add_argument("--com")
    ap.add_argument("--baud", type=int, default=9600)
    ap.add_argument("--replay")
    args = ap.parse_args()

    app = App(args.listen, args.port)

    if args.replay:
        app.stop_evt.clear()
        app.reader = ReaderThread(app.q, app.stop_evt, replay_path=args.replay)
        app.reader.start()
        app._set_status("Режим файла", True)
    elif args.com:
        app.port_var.set(args.com)
        app.baud_var.set(str(args.baud))
        app.connect()

    app.mainloop()


if __name__ == "__main__":
    main()
