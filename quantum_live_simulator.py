import argparse
import random
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

try:
    import serial
except Exception:
    serial = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


DIST_TO_CHECKPOINTS = {
    125: 1,
    250: 2,
    500: 4,
    1000: 8,
    2000: 16,
}


def _safe_bib(v) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    digits = "".join(ch for ch in s if ch.isdigit())
    if not digits:
        return ""
    n = int(digits)
    return "" if n <= 0 else str(n)


def load_bibs_from_xlsx(path: Path) -> List[str]:
    if load_workbook is None:
        raise RuntimeError("openpyxl не установлен")
    if not path.exists():
        raise FileNotFoundError(f"Не найден файл: {path}")
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    out: List[str] = []
    seen = set()
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        b = _safe_bib(row[0])
        if b and b not in seen:
            seen.add(b)
            out.append(b)
    if not out:
        raise RuntimeError("В xlsx не найдены номера в колонке A")
    return out


def fallback_bibs() -> List[str]:
    return ["1", "3", "5", "6", "11", "12", "13", "14", "30"]


def choose_bibs(pool: List[str], heat_idx: int, riders_per_heat: int) -> List[str]:
    if not pool:
        return []
    start = ((heat_idx - 1) * riders_per_heat) % len(pool)
    picked = [pool[(start + i) % len(pool)] for i in range(riders_per_heat)]
    uniq: List[str] = []
    seen = set()
    for b in picked:
        if b not in seen:
            seen.add(b)
            uniq.append(b)
    return uniq


def base_finish(distance: int) -> float:
    if distance == 125:
        return 3.75
    if distance == 250:
        return 7.55
    if distance == 500:
        return 15.95
    if distance == 1000:
        return 31.95
    if distance == 2000:
        return 63.95
    return max(3.75, distance / 30.0)


@dataclass
class TimedLine:
    offset_sec: float
    line: str


def build_heat_script(
    race: int,
    heat: int,
    bibs: List[str],
    checkpoints: int,
    distance: int,
    jitter: float,
) -> List[TimedLine]:
    now_str = datetime.now().strftime("%H:%M:%S.000")
    out: List[TimedLine] = []

    out.append(TimedLine(0.00, f"DN| {race}| {heat}|"))
    out.append(TimedLine(0.02, f"DA|  {race}| {heat}|" + "|".join(bibs)))
    out.append(TimedLine(0.05, f"DS|  {race}| {heat}|" + "|".join(bibs) + f"|{now_str}"))

    base = base_finish(distance)
    rider_finishes: List[Tuple[str, float]] = []
    for i, b in enumerate(bibs):
        bias = (i * 0.18) + random.uniform(-jitter, jitter)
        rider_finishes.append((b, max(0.8, base + bias)))

    events: List[TimedLine] = []
    for bib, f_time in rider_finishes:
        seg = f_time / float(checkpoints)
        for split_no in range(1, checkpoints):
            t = seg * split_no
            events.append(TimedLine(t, f"DI| {race}| {heat}| {split_no}|{bib}|      {t:.3f}|"))
        events.append(TimedLine(f_time, f"DF| {race}| {heat}| {checkpoints}|{bib}|   |      {f_time:.3f}|"))

    events.sort(key=lambda e: e.offset_sec)
    out.extend(events)
    return out


class Sink:
    def __init__(self, port: Optional[str], baud: int, stdout: bool):
        self.port = port
        self.baud = baud
        self.stdout = stdout
        self.ser = None

    def open(self):
        if self.port:
            if serial is None:
                raise RuntimeError("pyserial не установлен")
            self.ser = serial.Serial(
                self.port,
                self.baud,
                timeout=0.2,
                bytesize=serial.EIGHTBITS,
                parity=serial.PARITY_NONE,
                stopbits=serial.STOPBITS_ONE,
                xonxoff=False,
                rtscts=False,
                dsrdtr=False,
            )

    def close(self):
        if self.ser is not None:
            try:
                self.ser.close()
            except Exception:
                pass
            self.ser = None

    def send(self, line: str):
        msg = (line.rstrip() + "\n")
        if self.stdout:
            print(msg, end="")
        if self.ser is not None:
            self.ser.write(msg.encode("ascii", errors="ignore"))
            self.ser.flush()


def run_simulator(args) -> int:
    if args.distance not in DIST_TO_CHECKPOINTS:
        print(f"Unsupported distance: {args.distance}", file=sys.stderr)
        return 2

    if not args.port and not args.stdout:
        print("Укажи --port COMx или включи --stdout", file=sys.stderr)
        return 2

    bibs: List[str]
    try:
        xlsx = Path(args.xlsx)
        if not xlsx.exists() and xlsx.name == "test2.xlsx":
            alt = Path("тест2.xlsx")
            if alt.exists():
                xlsx = alt
        bibs = load_bibs_from_xlsx(xlsx)
    except Exception as e:
        bibs = fallback_bibs()
        print(f"WARN: не удалось прочитать xlsx, использую fallback bibs ({e})")

    checkpoints = DIST_TO_CHECKPOINTS[args.distance]
    sink = Sink(port=args.port, baud=args.baud, stdout=args.stdout)

    try:
        sink.open()
        print(f"SIM started: distance={args.distance}m checkpoints={checkpoints} heats={args.heats} speed={args.speed}x")
        if args.port:
            print(f"COM: {args.port} @ {args.baud}")

        race = args.race
        loop_idx = 0
        while True:
            loop_idx += 1
            for heat in range(1, args.heats + 1):
                heat_bibs = choose_bibs(bibs, heat + (loop_idx - 1) * args.heats, args.riders_per_heat)
                script = build_heat_script(
                    race=race,
                    heat=heat,
                    bibs=heat_bibs,
                    checkpoints=checkpoints,
                    distance=args.distance,
                    jitter=args.jitter,
                )

                t0 = time.monotonic()
                for ev in script:
                    due = t0 + (ev.offset_sec / max(0.05, args.speed))
                    wait = due - time.monotonic()
                    if wait > 0:
                        time.sleep(wait)
                    sink.send(ev.line)

                if args.gap > 0:
                    time.sleep(args.gap / max(0.05, args.speed))

            if not args.loop:
                break
            race += 1

    except KeyboardInterrupt:
        print("\nSIM stopped by user")
        return 0
    except Exception as e:
        print(f"SIM error: {e}", file=sys.stderr)
        return 1
    finally:
        sink.close()

    print("SIM finished")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Мини-эмулятор Quantum оборудования (live DN/DA/DS/DI/DF)")
    ap.add_argument("--port", default="", help="COM-порт для отправки (например COM7)")
    ap.add_argument("--baud", type=int, default=9600, help="Скорость порта")
    ap.add_argument("--stdout", action="store_true", help="Дублировать сообщения в консоль")
    ap.add_argument("--xlsx", default="тест2.xlsx", help="XLSX с номерами в колонке A")
    ap.add_argument("--distance", type=int, default=500, choices=[125, 250, 500, 1000, 2000], help="Дистанция")
    ap.add_argument("--race", type=int, default=1, help="Номер гонки")
    ap.add_argument("--heats", type=int, default=10, help="Количество заездов")
    ap.add_argument("--riders-per-heat", type=int, default=2, choices=[1, 2], help="Участников в заезде")
    ap.add_argument("--speed", type=float, default=1.0, help="Ускорение времени (2.0 = в 2 раза быстрее)")
    ap.add_argument("--jitter", type=float, default=0.04, help="Случайный разброс финиша, сек")
    ap.add_argument("--gap", type=float, default=1.0, help="Пауза между заездами, сек")
    ap.add_argument("--loop", action="store_true", help="Крутить заезды по кругу")
    args = ap.parse_args()
    return run_simulator(args)


if __name__ == "__main__":
    raise SystemExit(main())
