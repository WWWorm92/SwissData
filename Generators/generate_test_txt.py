import argparse
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


DIST_TO_CHECKPOINTS: Dict[int, int] = {
    125: 1,
    250: 2,
    500: 4,
    1000: 8,
    2000: 16,
}


def _safe_bib(v) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    digs = "".join(ch for ch in s if ch.isdigit())
    if not digs:
        return ""
    n = int(digs)
    return "" if n <= 0 else str(n)


def load_bibs(xlsx_path: Path) -> List[str]:
    if load_workbook is None:
        raise RuntimeError("openpyxl не установлен")
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Не найден файл: {xlsx_path}")

    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    out: List[str] = []
    seen = set()
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
        bib = _safe_bib(row[0])
        if bib and bib not in seen:
            seen.add(bib)
            out.append(bib)
    if not out:
        raise RuntimeError("В xlsx не найдены номера в колонке A")
    return out


def fallback_bibs() -> List[str]:
    # резервный набор из ранее используемого test2-пула
    return ["1", "3", "5", "6", "11", "12", "13", "14", "30"]


def choose_bibs_for_heat(all_bibs: List[str], heat_idx: int, riders_per_heat: int) -> List[str]:
    n = len(all_bibs)
    if n == 0:
        return []
    start = ((heat_idx - 1) * riders_per_heat) % n
    res = []
    for i in range(riders_per_heat):
        res.append(all_bibs[(start + i) % n])
    # remove duplicates while keeping order
    uniq = []
    seen = set()
    for b in res:
        if b not in seen:
            seen.add(b)
            uniq.append(b)
    return uniq


def base_finish_for_distance(distance_m: int) -> float:
    if distance_m == 125:
        return 3.75
    if distance_m == 250:
        return 7.55
    if distance_m == 500:
        return 15.95
    if distance_m == 1000:
        return 31.95
    if distance_m == 2000:
        return 63.95
    return max(3.75, distance_m / 30.0)


def make_heat_block(
    race: int,
    heat: int,
    bibs: List[str],
    checkpoints: int,
    start_dt: datetime,
    finish_base: float,
) -> List[str]:
    lines: List[str] = []
    lines.append(f"# Heat {race}-{heat}")
    lines.append(f"DN| {race}| {heat}|")
    if bibs:
        lines.append("DA|  {r}| {h}|{bs}".format(r=race, h=heat, bs="|".join(bibs)))
        lines.append(
            "DS|  {r}| {h}|{bs}|{st}".format(
                r=race,
                h=heat,
                bs="|".join(bibs),
                st=start_dt.strftime("%H:%M:%S.000"),
            )
        )
    else:
        lines.append(f"DA|  {race}| {heat}|")
        lines.append(f"DS|  {race}| {heat}|{start_dt.strftime('%H:%M:%S.000')}")

    for i_bib, bib in enumerate(bibs):
        bias = (i_bib * 0.18) + (heat % 5) * 0.02
        finish = finish_base + bias
        seg = finish / checkpoints
        for split_no in range(1, checkpoints):
            t_split = seg * split_no
            lines.append(f"DI| {race}| {heat}| {split_no}|{bib}|      {t_split:.3f}|")
        lines.append(f"DF| {race}| {heat}| {checkpoints}|{bib}|   |      {finish:.3f}|")

    lines.append("")
    return lines


def render_single_distance(
    distance_m: int,
    race: int,
    heats: int,
    bibs: List[str],
    riders_per_heat: int,
    start_hour: int,
) -> str:
    checkpoints = DIST_TO_CHECKPOINTS[distance_m]
    lines = [
        f"Quantum old format test data: only {distance_m}m ({checkpoints} checkpoints)",
        "Source bibs: тест2.xlsx",
        "",
    ]
    dt0 = datetime(2026, 1, 1, start_hour, 0, 0)
    finish_base = base_finish_for_distance(distance_m)
    for heat in range(1, heats + 1):
        hbibs = choose_bibs_for_heat(bibs, heat, riders_per_heat)
        lines.extend(
            make_heat_block(
                race=race,
                heat=heat,
                bibs=hbibs,
                checkpoints=checkpoints,
                start_dt=dt0 + timedelta(minutes=3 * (heat - 1)),
                finish_base=finish_base,
            )
        )
    return "\n".join(lines).rstrip() + "\n"


def render_mixed(
    race: int,
    bibs: List[str],
    riders_per_heat: int,
    start_hour: int,
) -> str:
    lines = [
        "Quantum old format test data: mixed run types in one file",
        "Source bibs: тест2.xlsx",
        "",
    ]
    dt0 = datetime(2026, 1, 1, start_hour, 0, 0)
    for idx, distance_m in enumerate([125, 250, 500, 1000, 2000], start=1):
        checkpoints = DIST_TO_CHECKPOINTS[distance_m]
        lines.append(f"# Heat {race}-{idx}: {distance_m}m ({checkpoints} checkpoints)")
        hbibs = choose_bibs_for_heat(bibs, idx, riders_per_heat)
        block = make_heat_block(
            race=race,
            heat=idx,
            bibs=hbibs,
            checkpoints=checkpoints,
            start_dt=dt0 + timedelta(minutes=3 * (idx - 1)),
            finish_base=base_finish_for_distance(distance_m),
        )
        # block already starts with # Heat line; skip first duplicated comment
        lines.extend(block[1:])
    return "\n".join(lines).rstrip() + "\n"


def main() -> None:
    ap = argparse.ArgumentParser(description="Генерация старых Quantum test txt (DN/DA/DS/DI/DF)")
    ap.add_argument("--xlsx", default="тест2.xlsx", help="Путь к xlsx с номерами (колонка A)")
    ap.add_argument("--out-dir", default="test_txt_by_distance", help="Папка вывода")
    ap.add_argument("--race", type=int, default=1, help="Номер гонки (race)")
    ap.add_argument("--heats", type=int, default=10, help="Кол-во заездов в каждом only_*.txt")
    ap.add_argument("--riders-per-heat", type=int, default=2, choices=[1, 2], help="Сколько участников в заезде")
    ap.add_argument("--skip-mixed", action="store_true", help="Не создавать mixed_types.txt")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists() and xlsx_path.name == "test2.xlsx":
        alt = Path("тест2.xlsx")
        if alt.exists():
            xlsx_path = alt

    try:
        bibs = load_bibs(xlsx_path)
    except Exception as e:
        bibs = fallback_bibs()
        print(f"WARN: не удалось прочитать {xlsx_path}: {e}")
        print("WARN: использую резервный набор номеров")
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    file_specs: List[Tuple[int, str, int]] = [
        (125, "only_125.txt", 10),
        (250, "only_250.txt", 11),
        (500, "only_500.txt", 12),
        (1000, "only_1000.txt", 13),
        (2000, "only_2000.txt", 14),
    ]

    for distance_m, fname, start_hour in file_specs:
        text = render_single_distance(
            distance_m=distance_m,
            race=args.race,
            heats=args.heats,
            bibs=bibs,
            riders_per_heat=args.riders_per_heat,
            start_hour=start_hour,
        )
        (out_dir / fname).write_text(text, encoding="utf-8", newline="\n")

    if not args.skip_mixed:
        mixed = render_mixed(
            race=args.race,
            bibs=bibs,
            riders_per_heat=args.riders_per_heat,
            start_hour=15,
        )
        (out_dir / "mixed_types.txt").write_text(mixed, encoding="utf-8", newline="\n")

    print(f"Готово: сгенерировано в {out_dir}")


if __name__ == "__main__":
    main()
